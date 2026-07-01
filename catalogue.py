"""
NDIS Support Catalogue loader.
Parses the NDIS Support Catalogue XLSX at startup and provides lookup functions.
"""

import os
from pathlib import Path
from openpyxl import load_workbook

CATALOGUE_PATH = Path(__file__).parent / "assets" / "NDIS Support Catalogue- 2026-27 v1.0.xlsx"
CATALOGUE_VERSION = "2026-27 v1.0"

_items: list[dict] = []

# From 2026-27 the catalogue uses a single National price limit; the previous
# per-state columns (ACT/NSW/NT/QLD/SA/TAS/VIC/WA) were removed on 1 July 2026.


def _load_catalogue() -> list[dict]:
    """Load and parse the Support Catalogue XLSX into a list of dicts."""
    if not CATALOGUE_PATH.exists():
        return []

    wb = load_workbook(CATALOGUE_PATH, read_only=True, data_only=True)
    ws = wb["Current Support Items"]

    headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        rows.append(record)

    wb.close()
    return rows


def get_items() -> list[dict]:
    """Return cached catalogue items, loading on first access."""
    global _items
    if not _items:
        _items = _load_catalogue()
    return _items


def search_items(query: str, state: str = "", limit: int = 20) -> list[dict]:
    """Search support items by item number or name keyword.

    Args:
        query: Support item number (e.g. '01_002_0107_1_1') or keyword(s) to search names
        state: Deprecated from 2026-27 — pricing is now National, so this no longer affects the price returned. Retained for backward compatibility.
        limit: Max results to return
    """
    items = get_items()
    query_lower = query.lower().strip()

    results = []
    for item in items:
        item_number = str(item.get("Support Item Number", "")).lower()
        item_name = str(item.get("Support Item Name", "")).lower()

        # Exact item number match
        if query_lower == item_number:
            results.append(item)
            continue

        # Partial item number match
        if query_lower.replace("_", "") in item_number.replace("_", ""):
            results.append(item)
            continue

        # Keyword search in name
        keywords = query_lower.split()
        if all(kw in item_name for kw in keywords):
            results.append(item)

    # Format results
    formatted = []
    for item in results[:limit]:
        entry = {
            "item_number": item.get("Support Item Number"),
            "name": item.get("Support Item Name"),
            "category": item.get("Support Category Name"),
            "registration_group": item.get("Registration Group Name"),
            "unit": item.get("Unit"),
            "quote_required": item.get("Quote"),
            "type": item.get("Type"),
        }

        # Add pricing — from 2026-27 the catalogue lists a single National price limit
        entry["price"] = item.get("National")
        entry["price_basis"] = "National"

        # Add remote pricing
        remote = item.get("Remote")
        very_remote = item.get("Very Remote")
        if remote is not None or very_remote is not None:
            entry["remote_price"] = remote
            entry["very_remote_price"] = very_remote

        # Add rules
        entry["non_face_to_face"] = item.get("Non-Face-to-Face Support Provision")
        entry["provider_travel"] = item.get("Provider Travel")
        entry["short_notice_cancellations"] = item.get("Short Notice Cancellations.")

        formatted.append(entry)

    return formatted
